"""SpreadsheetStream class."""

from __future__ import annotations

import typing as t
import re
import fnmatch
import csv
from openpyxl import load_workbook
from singer_sdk.streams import Stream
from singer_sdk import typing as th
from logging import getLogger
from decimal import Decimal
from datetime import datetime, date, time, timedelta, timezone
from concurrent.futures import ProcessPoolExecutor
from tap_spreadsheets.storage import Storage
import os

if t.TYPE_CHECKING:
    from singer_sdk.helpers.types import Context

logger = getLogger(__name__)


def _process_file_worker(args):
    """Worker to parse a file into records (picklable for ProcessPoolExecutor)."""
    cfg, file, headers, expected_types = args
    stream = cfg["stream"]

    rows = stream._iter_excel(file) if stream.format == "excel" else stream._iter_csv(file)
    records = []
    for row in rows:
        record = {
            h: stream._coerce_value(
                row[i + stream.skip_columns] if i + stream.skip_columns < len(row) else None,
                "integer" if "integer" in expected_types.get(h, []) else
                "number" if "number" in expected_types.get(h, []) else
                "string"
            )
            for i, h in enumerate(headers)
        }
        if stream.drop_empty and any(record.get(pk) in (None, "") for pk in stream.primary_keys):
            continue
        records.append(record)
    return file, records

class SpreadsheetStream(Stream):
    """Stream class for spreadsheet (CSV/Excel) files."""

    is_sorted = True
    
    def __init__(self, tap, file_cfg: dict) -> None:

        self.file_cfg = file_cfg
        self.path_glob: str = file_cfg["path"]
        self.format: str = file_cfg["format"].lower()
        self.worksheet_ref: str | int | None = file_cfg.get("worksheet")
        self.table_name = file_cfg.get("table_name")

        if not self.table_name:
            self.table_name = "spreadsheet_stream"
            if self.format == "excel" and self.worksheet_ref:
                self.table_name = f"{self.table_name}_{self.worksheet_ref}"

        super().__init__(tap, name=self.table_name)

        self.state_partitioning_keys = ["filename"]
        self.replication_key = "updated_at"
        self.forced_replication_method = "INCREMENTAL"

        self.primary_keys = [n.lower() for n in file_cfg.get("primary_keys", [])]
        self.drop_empty = file_cfg.get("drop_empty", True)
        self.skip_columns = file_cfg.get("skip_columns", 0)
        self.skip_rows = file_cfg.get("skip_rows", 0)
        self.sample_rows = file_cfg.get("sample_rows", 100)
        self.column_headers = file_cfg.get("column_headers")

        self._schema = None
        self._headers: list[str] = []
        self.storage = Storage(self.path_glob)


    def _stem_header(self, h: t.Any, idx: int) -> str:
        """Normalize header names to safe identifiers."""
        # If we accidentally get a Cell or other object, extract its value
        if hasattr(h, "value"):
            h = h.value

        if h is None or str(h).strip() == "":
            return f"col_{idx}"

        h = str(h)
        import unicodedata, re
        h = unicodedata.normalize("NFKD", h).encode("ascii", "ignore").decode()
        h = h.replace("\n", " ").replace("/", " ")
        h = h.lower()
        h = re.sub(r"[^a-z0-9]+", "_", h)
        h = re.sub(r"_+", "_", h)
        h = h.strip("_")

        return h or f"col_{idx}"

    def _infer_type(self, col_values: list[t.Any]):
        """Infer a JSON schema type from sample values with safe fallback."""
        if not col_values:
            return th.StringType()

        norm = []
        for v in col_values:
            if isinstance(v, Decimal):
                norm.append(float(v))
            else:
                norm.append(v)
        col_values = norm

        if all(isinstance(v, int) for v in col_values):
            return th.IntegerType()
        if all(isinstance(v, (int, float)) for v in col_values):
            return th.NumberType()
        if all(isinstance(v, str) for v in col_values):
            return th.StringType()
        if any(isinstance(v, (int, float)) for v in col_values):
            return th.NumberType()
        return th.StringType()

    def _coerce_value(self, v: t.Any, expected_type: str | None = None) -> t.Any:
        if v is None:
            return None
        if expected_type == "integer":
            try:
                return int(v)
            except Exception:
                return None
        if expected_type == "number":
            try:
                return float(v)
            except Exception:
                return None
        if expected_type == "string":
            return str(v)  
        if isinstance(v, Decimal):
            return float(v)
        if isinstance(v, (datetime, date, time)):
            return v.isoformat()
        if isinstance(v, timedelta):
            return v.total_seconds()
        return v


    def _extract_headers_excel(self, file: str) -> list[str]:
        with self.storage.open(file, "rb") as fh:
            wb = load_workbook(fh, read_only=True, data_only=True)

            ws = None
            if isinstance(self.worksheet_ref, int):
                if self.worksheet_ref < len(wb.worksheets):
                    ws = wb.worksheets[self.worksheet_ref]
            elif isinstance(self.worksheet_ref, str):
                if self.worksheet_ref in wb.sheetnames:
                    ws = wb[self.worksheet_ref]
                else:
                    pattern = self.worksheet_ref
                    regex = (
                        re.compile(fnmatch.translate(pattern))
                        if any(ch in pattern for ch in ["*", "?"])
                        else re.compile(pattern)
                    )
                    matches = [name for name in wb.sheetnames if regex.match(name)]
                    if matches:
                        ws = wb[matches[0]]
            if ws is None:
                logger.warning("No matching worksheet found in %s. Skipping file.", file)
                return []  # skip schema for this file

            header_row = ws.iter_rows(
                min_row=self.skip_rows + 1,
                max_row=self.skip_rows + 1,
                values_only=True,
            )
            raw_headers = next(header_row)
            return [self._stem_header(h, i) for i, h in enumerate(raw_headers)][self.skip_columns :]


    def _iter_excel(self, file: str):
        """Iterate data rows (excluding header) from all matched worksheets."""
        with self.storage.open(file, "rb") as fh:
            wb = load_workbook(fh, read_only=True, data_only=True)

            worksheets = []
            if isinstance(self.worksheet_ref, int):
                try:
                    worksheets = [wb.worksheets[self.worksheet_ref]]
                except IndexError:
                    logger.warning(
                        "Worksheet index %s out of range in %s. Skipping file.",
                        self.worksheet_ref, file,
                    )
                    return  # skip file

            elif isinstance(self.worksheet_ref, str):
                if self.worksheet_ref in wb.sheetnames:
                    worksheets = [wb[self.worksheet_ref]]
                else:
                    pattern = self.worksheet_ref
                    regex = (
                        re.compile(fnmatch.translate(pattern))
                        if any(ch in pattern for ch in ["*", "?"])
                        else re.compile(pattern)
                    )
                    matches = [name for name in wb.sheetnames if regex.match(name)]
                    if not matches:
                        logger.warning(
                            "No worksheets match '%s' in %s. Skipping file. Available: %s",
                            pattern, file, wb.sheetnames,
                        )
                        return  # skip file
                    worksheets = [wb[name] for name in matches]
            else:
                logger.warning("Invalid worksheet_ref %s. Skipping file %s", self.worksheet_ref, file)
                return

            for ws in worksheets:
                start_row = self.skip_rows + 2
                for row in ws.iter_rows(min_row=start_row, values_only=True):
                    yield row


    def _extract_headers_csv(self, file: str) -> list[str]:
        """Extract and normalize headers from a CSV file."""
        with self.storage.open(file, "rt") as fh:
            reader = csv.reader(fh)

            # Skip configured rows before header
            for _ in range(self.skip_rows):
                next(reader, None)

            try:
                raw_headers = next(reader)
            except StopIteration:
                raise ValueError(f"No header row found in {file}")

        headers: list[str] = []
        for i, h in enumerate(raw_headers):
            # Fallback if header is empty or looks numeric
            if h is None or str(h).strip() == "":
                headers.append(f"col_{i}")
            elif str(h).strip().isnumeric():
                headers.append(f"col_{i}")
            else:
                headers.append(self._stem_header(h, i))

        return headers[self.skip_columns :]

    def _iter_csv(self, file: str):
        with self.storage.open(file, "rt") as fh:
            sample = fh.read(4096)
            fh.seek(0)

            delimiter = self.file_cfg.get("delimiter")
            quotechar = self.file_cfg.get("quotechar")

            if not delimiter or not quotechar:
                try:
                    sniffer = csv.Sniffer()
                    dialect = sniffer.sniff(sample)
                    delimiter = delimiter or dialect.delimiter
                    quotechar = quotechar or dialect.quotechar
                except csv.Error:
                    delimiter = delimiter or ","
                    quotechar = quotechar or '"'

            reader = csv.reader(fh, delimiter=delimiter, quotechar=quotechar)

            # skip configured number of rows + the header
            for _ in range(self.skip_rows + 1):
                next(reader, None)

            for row in reader:
                yield row

    def _get_files(self) -> list[str]:
        files = self.storage.glob()
        if not files:
            raise FileNotFoundError(f"No {self.format} files found for {self.path_glob}")
        return files

    @property
    def schema(self) -> dict:
        if self._schema:
            return self._schema

        sample_file = self._get_files()[0]
        rows = (
            self._iter_excel(sample_file)
            if self.format == "excel"
            else self._iter_csv(sample_file)
        )

        if self.column_headers:
            headers = [self._stem_header(h, i) for i, h in enumerate(self.column_headers)]
        else:
            if self.format == "excel":
                headers = self._extract_headers_excel(sample_file)
            else:
                headers = self._extract_headers_csv(sample_file)

        self._headers = headers


        samples = []
        for i, row in enumerate(rows):
            if i >= self.sample_rows:
                break
            samples.append(row)

        types: dict[str, th.JSONTypeHelper] = {}
        for i, h in enumerate(headers):
            col_idx = i + self.skip_columns
            col_values = [
                r[col_idx] for r in samples if col_idx < len(r) and r[col_idx] not in (None, "")
            ]
            types[h] = self._infer_type(col_values)

        self._schema = th.PropertiesList(
            *(th.Property(name.lower(), tpe) for name, tpe in types.items()),
            th.Property(str(self.replication_key), th.DateTimeType(nullable=True), description="Updated date"),
        ).to_dict()

        overrides = self.file_cfg.get("schema_overrides", {})
        for col, props in overrides.items():
            if col in self._schema["properties"]:
                # Fix: replace Python None with the JSON Schema string "null"
                if "type" in props:
                    props["type"] = [
                        "null" if t is None else t
                        for t in props["type"]
                    ]
                self._schema["properties"][col].update(props)

        return self._schema

    def post_process(self, row: dict, context: Context | None = None) -> dict | None:
        """Post-Process, i.e. base64 decode the XML."""
        return {
            **row,
            str(self.replication_key): datetime.now(),
        }


    def get_records(self, context: Context | None) -> t.Iterable[dict]:
        if not self._headers:
            _ = self.schema
        headers = self._headers

        files = sorted(self._get_files())
        parallelize = max(1, self.file_cfg.get("parallelize", 1))

        expected_types = {
            name: schema_def.get("type", ["string"])
            for name, schema_def in self.schema["properties"].items()
        }

        def process_file(f: str):
            # file partition state
            partition_context = {"filename": os.path.basename(f)}
            last_bookmark = self.get_starting_replication_key_value(partition_context)

            mtime = datetime.fromtimestamp(os.path.getmtime(f), tz=timezone.utc)

            # --- skip file if already processed
            if last_bookmark:
                bookmark_dt = datetime.fromisoformat(last_bookmark)
                if bookmark_dt.tzinfo is None:
                    bookmark_dt = bookmark_dt.replace(tzinfo=timezone.utc)
                if mtime <= bookmark_dt:
                    logger.info("Skipping %s (mtime=%s <= bookmark=%s)", f, mtime, bookmark_dt)
                    return []

            # process rows
            _, recs = _process_file_worker(({"stream": self}, f, headers, expected_types))
            logger.info("Syncing %s (%d records)", f, len(recs))

            # attach replication key (per record)
            for r in recs:
                r[self.replication_key] = mtime

            # advance bookmark for this file
            if recs:
                self._increment_stream_state(
                    {"updated_at": mtime.isoformat()},
                    context=partition_context,
                )

            return recs

        if parallelize == 1:
            for f in files:
                yield from process_file(f)
        else:
            with ProcessPoolExecutor(max_workers=parallelize) as ex:
                results = list(ex.map(process_file, files))
                for recs in results:
                    yield from recs